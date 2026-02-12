from __future__ import annotations

import io
from datetime import datetime
from typing import Any


def export_l2_sections_to_xlsx_bytes(
    *,
    sections: list[dict[str, Any]] | None,
    title: str | None = None,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.worksheet.page import PageMargins
    except Exception as e:
        raise RuntimeError(f"缺少依赖 openpyxl（请安装: pip install openpyxl）。{repr(e)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "视频脚本"
    ws.sheet_view.showGridLines = False

    columns = [
        "段落",
        "镜头标题",
        "时长(秒)",
        "景别",
        "运镜",
        "场景",
        "道具",
        "画面",
        "字幕",
        "口播/音效",
        "BGM",
        "转场",
        "合规备注",
    ]

    total_cols = len(columns)

    thin = Side(style="thin", color="C9C9C9")
    medium = Side(style="medium", color="4F4F4F")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    section_fill = PatternFill("solid", fgColor="E6E6E6")
    section_font = Font(bold=True, color="1A1A1A", size=12)

    stripe_fill = PatternFill("solid", fgColor="FAFAFA")

    cell_align = Alignment(vertical="top", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    body_font = Font(size=11, color="111111")

    # Page / print settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    # Title row
    if title:
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.row_dimensions[1].height = 34
        c = ws.cell(row=1, column=1)
        c.font = Font(bold=True, size=18, color="111111")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        ws.row_dimensions[2].height = 18
        c = ws.cell(row=2, column=1)
        c.font = Font(color="666666", size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")

        start_row = 3
    else:
        start_row = 1

    # Header
    ws.append(columns)
    header_row = start_row
    ws.row_dimensions[header_row].height = 22
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_medium

    ws.freeze_panes = f"A{header_row + 1}"
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Column widths
    widths = [
        28,  # 段落
        18,  # 镜头标题
        10,  # 时长
        10,  # 景别
        12,  # 运镜
        18,  # 场景
        18,  # 道具
        48,  # 画面
        18,  # 字幕
        32,  # 口播/音效
        18,  # BGM
        12,  # 转场
        24,  # 合规备注
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    # Body
    row = header_row + 1
    sections = sections or []
    for sec in sections:
        if not isinstance(sec, dict):
            continue

        # Visual separator before each section (except first)
        if row > header_row + 1:
            ws.append([""] * total_cols)
            ws.row_dimensions[row].height = 10
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).border = Border(bottom=medium)
            row += 1

        sec_title = str(sec.get("section") or "")
        sec_rationale = str(sec.get("rationale") or "")
        sec_header = sec_title
        if sec_rationale.strip():
            sec_header = f"{sec_title}\n（设计理由）{sec_rationale}".strip()

        # Section merged row
        ws.append([sec_header] + [""] * (total_cols - 1))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.row_dimensions[row].height = 52
        for col in range(1, total_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill
            c.font = section_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border_medium
        row += 1

        sub = list(sec.get("sub_sections") or [])
        if not sub:
            ws.append(["（无镜头）"] + [""] * (total_cols - 1))
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = body_font
                cell.alignment = cell_align
                cell.border = border_medium
            ws.row_dimensions[row].height = 18
            row += 1
            continue

        for k, seg in enumerate(sub):
            if not isinstance(seg, dict):
                continue

            props = seg.get("props")
            if isinstance(props, list):
                props_str = "、".join([str(x) for x in props if x is not None])
            else:
                props_str = str(props or "")

            ws.append(
                [
                    "",  # 段落列留空（更像脚本排版）
                    str(seg.get("title") or ""),
                    seg.get("duration_s"),
                    str(seg.get("shot") or ""),
                    str(seg.get("camera_move") or ""),
                    str(seg.get("location") or ""),
                    props_str,
                    str(seg.get("visual") or ""),
                    str(seg.get("onscreen_text") or ""),
                    str(seg.get("audio") or ""),
                    str(seg.get("music") or ""),
                    str(seg.get("transition") or ""),
                    str(seg.get("compliance_notes") or ""),
                ]
            )

            is_stripe = (k % 2) == 1
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = body_font
                # Center some columns for readability
                if col in (3, 4, 5, 12):
                    cell.alignment = cell_align_center
                else:
                    cell.alignment = cell_align

                # Zebra stripe for shot rows
                if is_stripe:
                    cell.fill = stripe_fill

                # Stronger border at section boundaries
                if k == 0:
                    cell.border = Border(
                        left=medium,
                        right=medium,
                        top=medium,
                        bottom=thin,
                    )
                elif k == len(sub) - 1:
                    cell.border = Border(
                        left=medium,
                        right=medium,
                        top=thin,
                        bottom=medium,
                    )
                else:
                    cell.border = Border(
                        left=medium,
                        right=medium,
                        top=thin,
                        bottom=thin,
                    )

            ws.row_dimensions[row].height = 78
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
